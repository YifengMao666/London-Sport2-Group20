import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

CODEBOOK_FILES = {
    1: "8223_code_book.xlsx",
    2: "8391_active_lives_adult_codebook.xlsx",
    3: "8651_active_lives_adult_code_book_full_year_3_nov_17-18_final.xlsx",
    4: "8652_active_lives_adult_code_book_full_year_4.xlsx",
    5: "8899_active_lives_adult_code_book_full_year_5_2019-20.xlsx",
    6: "8993_active_lives_survey_full_year_6_code_book_final.xlsx",
    7: "9136_active_lives_survey_year_7_code_book.xlsx",
    8: "9288_active_lives_survey_year_8_code_book.xlsx",
}

CODEBOOK_DIR = Path(r"C:\Users\Lenovo\Desktop\Dissertation\Data\8_codebook")
OUTPUT_PATH = CODEBOOK_DIR / "125_activities_composites_year1_to_year8_REBUILT.xlsx"

ALL_YEARS = list(range(1, 9))

def normalise_item_codes(raw_string):
    """Turn a raw 'Constituent activities' cell into a sorted tuple of
    upper-case, standardised item codes."""
    if raw_string is None:
        return tuple()

    cleaned = raw_string.replace(",", " ")
    tokens = [tok.strip() for tok in cleaned.split() if tok.strip()]

    normalised = []
    for tok in tokens:
        tok = tok.upper()
        tok = re.sub(r"^ACTIV1_", "A1_", tok)
        tok = re.sub(r"^ACTIV2_", "A2_", tok)
        normalised.append(tok)

    return tuple(sorted(set(normalised)))


def read_year_composites(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Activity composites"]

    result = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        dv_suffix = row[0]
        if dv_suffix is None or str(dv_suffix).strip() == "":
            continue
        dv_suffix = str(dv_suffix).strip()
        activity_name = row[1]
        raw_items = row[2]
        result[dv_suffix] = (activity_name, normalise_item_codes(raw_items))

    wb.close()
    return result


def describe_changes(year_to_items):
    parts = []
    prev_year = ALL_YEARS[0]
    prev_items = set(year_to_items[prev_year])

    for year in ALL_YEARS[1:]:
        current_items = set(year_to_items[year])
        added = sorted(current_items - prev_items)
        removed = sorted(prev_items - current_items)

        if added or removed:
            bits = []
            if added:
                bits.append("added " + ", ".join(added))
            if removed:
                bits.append("removed " + ", ".join(removed))
            parts.append(f"From Y{year}: " + " and ".join(bits))

        prev_year = year
        prev_items = current_items

    return "; ".join(parts)

def load_all_years():
    year_data = {}
    for year, filename in CODEBOOK_FILES.items():
        path = CODEBOOK_DIR / filename
        print(f"Reading Year {year}: {filename}")
        year_data[year] = read_year_composites(path)
    return year_data

def classify_activities(year_data):
    all_suffixes = set()
    for year_dict in year_data.values():
        all_suffixes.update(year_dict.keys())

    stable_rows = []
    changed_rows = []
    not_stable_rows = []

    for suffix in sorted(all_suffixes):
        years_present = [y for y in ALL_YEARS if suffix in year_data[y]]

        first_year = years_present[0]
        activity_name = year_data[first_year][suffix][0]

        if len(years_present) < len(ALL_YEARS):
            missing_years = [y for y in ALL_YEARS if y not in years_present]
            flags = {y: ("Y" if y in years_present else "N") for y in ALL_YEARS}
            not_stable_rows.append(
                (suffix, activity_name, flags, missing_years)
            )
            continue

        year_to_items = {y: year_data[y][suffix][1] for y in ALL_YEARS}
        item_sets = {year_to_items[y] for y in ALL_YEARS}

        if len(item_sets) == 1:
            item_count = len(year_to_items[ALL_YEARS[0]])
            definition_str = ", ".join(year_to_items[ALL_YEARS[0]])
            stable_rows.append(
                (suffix, activity_name, item_count, definition_str)
            )
        else:
            counts = {y: len(year_to_items[y]) for y in ALL_YEARS}
            change_desc = describe_changes(year_to_items)
            changed_rows.append(
                (suffix, activity_name, counts, change_desc)
            )

    return stable_rows, not_stable_rows, changed_rows


def write_workbook(stable_rows, not_stable_rows, changed_rows, output_path):
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    ws1 = wb.active
    ws1.title = "1_Stable composites"
    ws1["A1"] = (
        f"Activity Composites that exist in all eight years (Y1-Y8) "
        f"with fully consistent definitions ({len(stable_rows)} in total)"
    )
    ws1["A2"] = (
        "Composite codes have been standardized "
        "(ACTIV1_/ACTIV2_ unified to A1_/A2_)."
    )
    header_row = 4
    headers1 = [
        "DV suffix", "Activity Name", "Number of Composite Items",
        "Composite Definition (consistent across eight years)",
    ]
    for col, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=header_row, column=col, value=h)
        cell.font = bold

    for i, (suffix, name, count, definition) in enumerate(stable_rows, start=1):
        r = header_row + i
        ws1.cell(row=r, column=1, value=suffix)
        ws1.cell(row=r, column=2, value=name)
        ws1.cell(row=r, column=3, value=count)
        ws1.cell(row=r, column=4, value=definition)

    ws2 = wb.create_sheet("2_Common but not stable")
    ws2["A1"] = (
        f"Activity Composites that exist in only some years "
        f"({len(not_stable_rows)} in total)"
    )
    header_row2 = 3
    headers2 = ["DV suffix", "Activity Name"] + [f"Y{y}" for y in ALL_YEARS] + ["Missing Years"]
    for col, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=header_row2, column=col, value=h)
        cell.font = bold

    for i, (suffix, name, flags, missing_years) in enumerate(not_stable_rows, start=1):
        r = header_row2 + i
        ws2.cell(row=r, column=1, value=suffix)
        ws2.cell(row=r, column=2, value=name)
        for j, y in enumerate(ALL_YEARS, start=3):
            ws2.cell(row=r, column=j, value=("\u2713" if flags[y] == "Y" else "\u2717"))
        missing_str = ", ".join(f"Y{y}" for y in missing_years)
        ws2.cell(row=r, column=3 + len(ALL_YEARS), value=missing_str)

    ws3 = wb.create_sheet("3_Composition changed")
    ws3["A1"] = (
        f"Activity Composites present in all eight years with unchanged "
        f"names, but with additions or removals in composite items "
        f"({len(changed_rows)} in total)"
    )
    ws3["A2"] = (
        "The number columns show the count of composite items for that "
        "year. The change description uses the earliest year as the "
        "baseline and describes additions/removals in chronological "
        "order. Note these definition changes when comparing across years."
    )
    header_row3 = 4
    headers3 = ["DV suffix", "Activity Name"] + [f"Y{y}" for y in ALL_YEARS] + ["Change Description"]
    for col, h in enumerate(headers3, start=1):
        cell = ws3.cell(row=header_row3, column=col, value=h)
        cell.font = bold

    for i, (suffix, name, counts, change_desc) in enumerate(changed_rows, start=1):
        r = header_row3 + i
        ws3.cell(row=r, column=1, value=suffix)
        ws3.cell(row=r, column=2, value=name)
        for j, y in enumerate(ALL_YEARS, start=3):
            ws3.cell(row=r, column=j, value=counts[y])
        ws3.cell(row=r, column=3 + len(ALL_YEARS), value=change_desc)

    wb.save(output_path)
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    year_data = load_all_years()
    stable_rows, not_stable_rows, changed_rows = classify_activities(year_data)

    print()
    print("Stable composites:      ", len(stable_rows))
    print("Common but not stable:  ", len(not_stable_rows))
    print("Composition changed:    ", len(changed_rows))
    print("Total DV suffixes:      ",
          len(stable_rows) + len(not_stable_rows) + len(changed_rows))

    write_workbook(stable_rows, not_stable_rows, changed_rows, OUTPUT_PATH)
